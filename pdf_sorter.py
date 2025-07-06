import os
import time
import json
import shutil
from datetime import datetime
from pdf2image import convert_from_path
import pytesseract
from openpyxl import Workbook, load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler


def load_config(path):
    with open(path, 'r') as f:
        config = json.load(f)
    config['input_dir'] = os.path.expanduser(config['input_dir'])
    config['belege_dir'] = os.path.expanduser(config['belege_dir'])
    config['excel_path'] = os.path.expanduser(config['excel_path'])
    return config


def extract_text(pdf_path):
    text = []
    images = convert_from_path(pdf_path)
    for img in images:
        text.append(pytesseract.image_to_string(img, lang='deu+eng'))
    return "\n".join(text).lower()


def classify(text, category_map):
    for keyword, folder in category_map.items():
        if keyword.lower() in text:
            return folder
    return None


def add_hyperlink(excel_path, file_path):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Datum", "Datei", "Pfad"])

    row = [datetime.now().strftime('%Y-%m-%d'), os.path.basename(file_path), file_path]
    ws.append(row)
    cell = ws.cell(row=ws.max_row, column=3)
    cell.hyperlink = file_path
    cell.style = "Hyperlink"
    wb.save(excel_path)


def process_file(pdf_path, config):
    text = extract_text(pdf_path)
    category = classify(text, config['category_map'])
    if category:
        dest_dir = os.path.join(config['belege_dir'], category)
    else:
        dest_dir = os.path.join(config['belege_dir'], 'Unsortiert')
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(dest_dir, os.path.basename(pdf_path))
    shutil.move(pdf_path, dest_path)
    add_hyperlink(config['excel_path'], dest_path)
    print(f"Processed {pdf_path} -> {dest_path}")


class PDFHandler(FileSystemEventHandler):
    def __init__(self, config):
        self.config = config

    def on_created(self, event):
        if not event.is_directory and event.src_path.lower().endswith('.pdf'):
            # give file some time to be written completely
            time.sleep(1)
            process_file(event.src_path, self.config)


def watch_folder(config):
    observer = Observer()
    handler = PDFHandler(config)
    observer.schedule(handler, path=config['input_dir'], recursive=False)
    observer.start()
    print(f"Watching {config['input_dir']} for PDFs...")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == '__main__':
    CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config.json')
    config = load_config(CONFIG_PATH)
    os.makedirs(config['input_dir'], exist_ok=True)
    watch_folder(config)
